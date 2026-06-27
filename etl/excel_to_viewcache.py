#!/usr/bin/env python3
"""
ETL: Excel → view_cache.json (MARQOOM_SCHEMA v1)
يقرأ أي ملف Excel من مشروع مرقوم ويولّد view_cache.json بنفس بنية البحر المحيط
"""

import json
import sys
import os
import re
from datetime import datetime, timezone
import openpyxl

# ─── خريطة أسماء الأوراق إلى مفاتيح JSON ───────────────────────────────────
SHEET_MAP = {
    # ورقة الملخص / البيانات الأساسية
    "بيانات الكتاب":       "book_data",
    "بيانات الملفات":      "book_data",
    "الملخص التنفيذي":     "executive_summary",
    "ملخص تنفيذي":         "executive_summary",
    "لوحة المعلومات":      "dashboard",
    "لوحة عامة":           "dashboard",
    "المؤشرات المركبة":    "composite_indicators",

    # الأوراق الرئيسية
    "المصطلحات المنهجية":  "terminology",
    "المصطلحات":           "terminology",
    "توزيع الأغراض":       "purposes",
    "الكتب":               "books",
    "الكتب المصرح بها":    "books",
    "الأعلام":             "people",
    "المدارس والاتجاهات":  "schools",
    "المدارس":             "schools",
    "الموارد حسب المجال":  "resources_by_field",
    "المجالات":            "resources_by_field",

    # أوراق الخلاف والترجيح والنقد
    "الخلاف":              "disagreement",
    "الترجيح":             "preponderance",
    "النقد":               "criticism",
    "الإحالات":            "references",
    "الإحالات الداخلية":   "references",
    "الخلاف والترجيح والنقد": "disagreement_combined",
    "الخلاف والترجيح":     "disagreement_combined",

    # أوراق متخصصة
    "صيغ الأداء والرواية": "narration_forms",
    "صيغ الأداء":          "narration_forms",
    "صيغ الرواية":         "narration_forms",
    "الجرح والتعديل":      "hadith_criticism",
    "الكثافة العلمية":     "density",
    "الكثافة":             "density",
    "الحكم الحديثي":       "hadith_ruling",
    "السور والصفحات":      "surah_distribution",
    "إحصاء الأجزاء":       "parts_stats",
    "الأجزاء":             "parts_stats",

    # أوراق المخرجات
    "حدود الدراسة":        "limits",
    "للمستفيد السريع":     "quick_summary",
    "نتائج قابلة للنشر":   "publishable_results",
    "مخرجات نشر":          "publishable_results",
    "أفكار أبحاث":         "research_ideas",
    "المراجعة اليدوية":    "manual_review",
    "ما يحتاج مراجعة":     "manual_review",
    "شواهد مختارة":        "selected_examples",
    "عينة الصفحات":        "page_sample",
    "الألفاظ الشائعة":     "common_terms",
    "الألفاظ الزمنية":     "temporal_terms",
    "اللغة والاستنباط":    "language_deduction",
    "التراجم والسيرة العلمية": "biographies",
    "تدخل المؤلف والتعقيب": "author_intervention",
    "فهرس الأنواع":        "types_index",
    "علوم القرآن":         "quran_sciences",
    "عناوين الطبقات":      "tabaqat_titles",
    "عينات التراجم":       "biography_samples",
    "عينات غير التراجم":   "non_biography_samples",
    "ملحق الموارد":        "resources_appendix",
    "المحذوفات المشتركة":  "common_exclusions",
    "المحذوفات":           "common_exclusions",
    "منهجية العمل":        "methodology",
    "تصور بصري للموقع":    None,  # تُتجاهل
    "تنبيهات منهجية":      None,  # تُتجاهل
    "عرض بصري":            None,  # تُتجاهل
}


def normalize_sheet_name(name: str) -> str:
    """تطبيع اسم الورقة بإزالة المسافات الزائدة"""
    return name.strip()


def sheet_to_key(name: str) -> str | None:
    """تحويل اسم الورقة إلى مفتاح JSON"""
    normalized = normalize_sheet_name(name)
    return SHEET_MAP.get(normalized, f"sheet_{normalized.replace(' ', '_')}")


def read_sheet_as_table(ws) -> dict:
    """قراءة ورقة Excel كجدول بأعمدة وصفوف"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"_columns": [], "rows": []}

    # الصف الأول هو العناوين
    headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]

    data_rows = []
    for row in rows[1:]:
        # تحويل القيم
        converted = []
        for v in row:
            if v is None:
                converted.append(None)
            elif isinstance(v, (int, float)):
                converted.append(v)
            else:
                converted.append(str(v).strip())
        # تخطي الصفوف الفارغة كلياً
        if all(v is None or v == "" for v in converted):
            continue
        data_rows.append(converted)

    return {"_columns": headers, "rows": data_rows}


def extract_book_metadata(wb: openpyxl.Workbook) -> dict:
    """استخراج بيانات الكتاب من أوراق البيانات"""
    meta = {
        "book": None,
        "author": None,
        "metrics": {}
    }

    # البحث في أوراق البيانات
    for sheet_name in ["بيانات الكتاب", "بيانات الملفات", "لوحة المعلومات", "لوحة عامة", "الملخص التنفيذي", "ملخص تنفيذي"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            val = row[1] if len(row) > 1 else None
            val_str = str(val).strip() if val is not None else None

            if not val_str or val_str == "None":
                continue

            # استخراج اسم الكتاب
            if any(k in key for k in ["الكتاب", "عنوان", "اسم المصنف", "المصنف"]):
                if not meta["book"]:
                    meta["book"] = val_str

            # استخراج اسم المؤلف
            if any(k in key for k in ["المؤلف", "المصنف", "صاحب", "الإمام"]):
                if not meta["author"]:
                    meta["author"] = val_str

            # استخراج المقاييس
            if any(k in key for k in ["الصفحات", "عدد الصفحات"]):
                try:
                    meta["metrics"]["pages"] = int(str(val).replace(",", "").replace("،", ""))
                except:
                    meta["metrics"]["pages"] = val_str

            if any(k in key for k in ["الكلمات", "عدد الكلمات", "الكلمات التقريبي"]):
                try:
                    meta["metrics"]["words_approx"] = int(str(val).replace(",", "").replace("،", ""))
                except:
                    meta["metrics"]["words_approx"] = val_str

            if any(k in key for k in ["الأجزاء", "عدد الأجزاء", "المجلدات"]):
                try:
                    meta["metrics"]["parts"] = int(str(val).replace(",", "").replace("،", ""))
                except:
                    meta["metrics"]["parts"] = val_str

            if any(k in key for k in ["الحروف", "عدد الحروف"]):
                try:
                    meta["metrics"]["chars"] = int(str(val).replace(",", "").replace("،", ""))
                except:
                    meta["metrics"]["chars"] = val_str

    return meta


def extract_top_list(ws, max_rows: int = 25) -> list:
    """استخراج قائمة أعلى العناصر من ورقة"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    result = []

    for row in rows[1:max_rows + 1]:
        if not row or all(v is None for v in row):
            continue
        converted = []
        for v in row:
            if v is None:
                converted.append(None)
            elif isinstance(v, (int, float)):
                converted.append(v)
            else:
                converted.append(str(v).strip())
        result.append(converted)

    return result


def excel_to_viewcache(excel_path: str, kashaf_id: str, book_title: str = None, author_name: str = None) -> dict:
    """التحويل الرئيسي: Excel → dict بنية view_cache"""

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheets = wb.sheetnames

    # استخراج البيانات الأساسية
    meta = extract_book_metadata(wb)

    # تجاوز القيم اليدوية إذا وُجدت
    if book_title:
        meta["book"] = book_title
    if author_name:
        meta["author"] = author_name

    # بناء الـ view_cache
    cache = {
        "_meta": {
            "schema_version": "MARQOOM_SCHEMA_v1",
            "book_id": kashaf_id,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "source": "excel_etl"
        },
        "project": "مرقوم — بوابة الكشافات التراثية الرقمية",
        "book": meta["book"] or os.path.basename(excel_path).replace(".xlsx", ""),
        "author": meta["author"] or "",
        "metrics": meta["metrics"],
        "tables": {}
    }

    # معالجة كل ورقة
    for sheet_name in sheets:
        json_key = sheet_to_key(sheet_name)
        if json_key is None:
            continue  # تُتجاهل

        ws = wb[sheet_name]
        table_data = read_sheet_as_table(ws)

        if not table_data["rows"]:
            continue

        # الأوراق الرئيسية تُضاف مباشرة كـ top_* للتوافق مع العارض القديم
        if json_key == "purposes":
            cache["top_purposes"] = extract_top_list(ws, 15)
        elif json_key == "books":
            cache["top_books"] = extract_top_list(ws, 30)
        elif json_key == "people":
            cache["top_people"] = extract_top_list(ws, 30)
        elif json_key == "resources_by_field":
            cache["top_resources"] = extract_top_list(ws, 15)
        elif json_key == "surah_distribution":
            cache["surah_distribution"] = extract_top_list(ws, 120)
        elif json_key == "limits":
            cache["limits"] = extract_top_list(ws, 20)
        else:
            # باقي الأوراق تُضاف في قسم tables
            cache["tables"][json_key] = {
                "title": sheet_name,
                **table_data
            }

    wb.close()

    # إزالة tables إذا كانت فارغة
    if not cache["tables"]:
        del cache["tables"]

    return cache


def main():
    if len(sys.argv) < 3:
        print("Usage: python excel_to_viewcache.py <excel_file> <kashaf_id> [book_title] [author_name]")
        print("Example: python excel_to_viewcache.py fathalbaari.xlsx fathalbaari 'فتح الباري' 'ابن حجر'")
        sys.exit(1)

    excel_path = sys.argv[1]
    kashaf_id = sys.argv[2]
    book_title = sys.argv[3] if len(sys.argv) > 3 else None
    author_name = sys.argv[4] if len(sys.argv) > 4 else None

    if not os.path.exists(excel_path):
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)

    print(f"Processing: {os.path.basename(excel_path)}")
    cache = excel_to_viewcache(excel_path, kashaf_id, book_title, author_name)

    output_path = f"/tmp/viewcache_{kashaf_id}.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

    print(f"Generated: {output_path}")
    print(f"  book: {cache['book']}")
    print(f"  author: {cache['author']}")
    print(f"  metrics: {cache['metrics']}")
    print(f"  top_purposes: {len(cache.get('top_purposes', []))} items")
    print(f"  top_books: {len(cache.get('top_books', []))} items")
    print(f"  top_people: {len(cache.get('top_people', []))} items")
    print(f"  tables: {list(cache.get('tables', {}).keys())}")

    return output_path


if __name__ == "__main__":
    main()
