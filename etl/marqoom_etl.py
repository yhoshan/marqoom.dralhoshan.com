#!/usr/bin/env python3
"""
marqoom_etl.py — محوّل مرقوم العام
الإصدار: 1.0.0
الغرض: تحويل أي ملف Excel ملتزم ببروتوكول مرقوم إلى:
  - view_cache.json  (View Cache للعرض)
  - validation_report.json (تقرير التحقق)

الاستخدام:
  python3 marqoom_etl.py <excel_file> --book-id marqoom-12 --title "صحيح البخاري" \
      --author "البخاري" --died-hijri 256 --category حديث [--output-dir ./output]
"""

import argparse
import hashlib
import json
import os
import re
import sys
from datetime import datetime, timezone
from typing import Any, Optional

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl غير مثبت. شغّل: pip3 install openpyxl")
    sys.exit(1)

# ============================================================
# Schema Registry — مستخرج من MARQOOM_PROTOCOL.md
# ============================================================
SCHEMA_REGISTRY = {
    "core": {
        "لوحة عامة": {
            "aliases": ["لوحة المعلومات"],
            "required": True,
            "key_columns": ["المؤشر", "القيمة"],
            "json_key": "dashboard",
        },
        "بيانات الملفات": {
            "aliases": ["بيانات الكتاب"],
            "required": True,
            "key_columns": ["اسم الملف", "عدد الكلمات", "عدد الحروف"],
            "json_key": "file_data",
        },
        "توزيع الأغراض": {
            "aliases": [],
            "required": True,
            "key_columns": ["الغرض", "العدد", "النسبة"],
            "json_key": "purposes",
        },
        "الموارد حسب المجال": {
            "aliases": ["المجالات", "مجالات الموارد", "موارد حسب المجال"],
            "required": True,
            "key_columns": ["المجال", "العدد"],
            "json_key": "resources_by_domain",
        },
        "حدود الدراسة": {
            "aliases": [],
            "required": True,
            "key_columns": ["الحد/التنبيه", "البند"],
            "json_key": "limits",
        },
        "المؤشرات المركبة": {
            "aliases": ["المؤشرات"],
            "required": True,
            "key_columns": ["المؤشر", "النسبة"],
            "json_key": "composite_indicators",
        },
    },
    "modules": {
        "ملخص تنفيذي": {
            "aliases": ["الملخص التنفيذي", "نتائج قابلة للنشر", "نتائج للنشر", "نتائج النشر", "مخرجات نشر", "ملخص"],
            "json_key": "executive_summary",
        },
        "المصطلحات المنهجية": {
            "aliases": ["المصطلحات", "مصطلحات منهجية", "مصطلحات"],
            "json_key": "terms",
        },
        "الكتب المصرح بها": {
            "aliases": ["الكتب", "كتب الموارد", "الموارد-الكتب", "كثافة الكتب", "كتب مصرح بها", "كتب"],
            "json_key": "books",
        },
        "الأعلام": {
            "aliases": ["الأعلام والمصادر البشرية", "الموارد-الأعلام"],
            "json_key": "people",
        },
        "المدارس والاتجاهات": {
            "aliases": ["المدارس", "المدارس والجهات العلمية", "مدارس"],
            "json_key": "schools",
        },
        "صيغ الأداء والرواية": {
            "aliases": ["صيغ الأداء", "صيغ الرواية"],
            "json_key": "narration_forms",
        },
        "الخلاف والترجيح": {
            "aliases": ["الخلاف والترجيح والنقد", "الخلاف", "الترجيح", "النقد", "خلاف ترجيح نقد", "خلاف وترجيح"],
            "json_key": "disagreement",
        },
        "المحذوفات المشتركة": {
            "aliases": ["المحذوفات", "محذوفات مشتركة", "محذوفات"],
            "json_key": "exclusions",
        },
        "السور والصفحات": {
            "aliases": ["السور"],
            "json_key": "surahs",
        },
        "الكثافة العلمية": {
            "aliases": ["الكثافة", "الكثافة حسب الكتب"],
            "json_key": "density",
        },
        "الإحالات الداخلية": {
            "aliases": ["الإحالات"],
            "json_key": "internal_refs",
        },
        "للمستفيد السريع": {
            "aliases": ["المستفيد السريع", "مستفيد سريع"],
            "json_key": "quick_reader",
        },
        "أفكار أبحاث": {
            "aliases": ["أفكار للبحث", "أفكار"],
            "json_key": "research_ideas",
        },
        "منهجية العمل": {
            "aliases": ["منهجية", "منهجية العمل"],
            "json_key": "methodology",
        },
        "الحكم الحديثي": {
            "aliases": [],
            "json_key": "hadith_rulings",
        },
        "علوم القرآن": {
            "aliases": ["فهرس الأنواع"],
            "json_key": "quran_sciences",
        },
    },
}

# ============================================================
# أدوات مساعدة
# ============================================================

def normalize_text(text: str) -> str:
    """تطبيع النص: إزالة مسافات زائدة وتوحيد الأحرف"""
    if not text:
        return ""
    text = str(text).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_col(col: str) -> str:
    """تطبيع اسم العمود للمقارنة"""
    col = normalize_text(col)
    col = col.replace("%", "").replace("٪", "").strip()
    col = re.sub(r"^\d+[\.\-\s]+", "", col)  # إزالة أرقام البداية
    return col


def fuzzy_match_sheet(sheet_names: list[str], target: str, aliases: list[str]) -> Optional[str]:
    """إيجاد ورقة Excel بمطابقة مرنة"""
    all_targets = [target] + aliases
    # مطابقة تامة أولاً
    for t in all_targets:
        for s in sheet_names:
            if normalize_text(s) == normalize_text(t):
                return s
    # مطابقة جزئية
    for t in all_targets:
        for s in sheet_names:
            if normalize_text(t) in normalize_text(s) or normalize_text(s) in normalize_text(t):
                return s
    return None


def to_number(val: Any) -> Optional[float]:
    """تحويل قيمة إلى رقم"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("٪", "").replace("%", "")
    # تحويل الأرقام العربية
    arabic_digits = "٠١٢٣٤٥٦٧٨٩"
    for i, d in enumerate(arabic_digits):
        s = s.replace(d, str(i))
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def read_sheet_as_table(ws) -> dict:
    """قراءة ورقة Excel كجدول مع _columns و rows"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"_columns": [], "rows": []}

    # إيجاد أول صف غير فارغ كعناوين
    header_row = None
    data_start = 0
    for i, row in enumerate(rows):
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if len(non_empty) >= 2:
            header_row = row
            data_start = i + 1
            break

    if header_row is None:
        return {"_columns": [], "rows": []}

    columns = [normalize_text(str(c)) if c is not None else "" for c in header_row]
    # إزالة الأعمدة الفارغة من النهاية
    while columns and not columns[-1]:
        columns.pop()

    result_rows = []
    for row in rows[data_start:]:
        row_data = list(row[: len(columns)])
        # تجاهل الصفوف الفارغة كلياً
        if all(c is None or str(c).strip() == "" for c in row_data):
            continue
        # تطبيع القيم
        normalized = []
        for c in row_data:
            if c is None:
                normalized.append(None)
            elif isinstance(c, (int, float)):
                normalized.append(c)
            else:
                normalized.append(normalize_text(str(c)))
        result_rows.append(normalized)

    return {"_columns": columns, "rows": result_rows}


def extract_summary_from_dashboard(ws) -> dict:
    """استخراج الأرقام الكبرى من لوحة عامة"""
    summary = {
        "pages": None,
        "parts": None,
        "words_approx": None,
        "chars": None,
        "surahs_detected": None,
    }
    KEYWORDS = {
        "pages": ["صفحة", "الصفحات", "عدد الصفحات", "ملفات xhtml", "ملف xhtml"],
        "parts": ["جزء", "الأجزاء", "عدد الأجزاء"],
        "words_approx": ["كلمة", "الكلمات", "عدد الكلمات", "كلمات المتن", "كلمات المتن المعتمدة"],
        "chars": ["حرف", "الحروف", "عدد الحروف"],
        "surahs_detected": ["سورة", "السور", "عدد السور"],
    }
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        label = normalize_text(str(row[0])).lower()
        value = row[1] if len(row) > 1 else None
        for key, keywords in KEYWORDS.items():
            if any(kw.lower() in label for kw in keywords):
                # محاولة استخراج رقم من القيمة أو من نص المؤشر نفسه
                num = to_number(value)
                if num is None and value:
                    # استخراج أول رقم من النص
                    val_str = str(value).replace(",", "")
                    m = re.search(r"[\d.]+", val_str)
                    if m:
                        num = to_number(m.group())
                if num is not None and num > 100:  # تجاهل نسب مئوية صغيرة
                    summary[key] = int(num) if num == int(num) else num
    return summary


def extract_top_n(ws, label_col_hints: list[str], count_col_hints: list[str], n: int = 10) -> dict:
    """استخراج أعلى N عنصراً من ورقة"""
    table = read_sheet_as_table(ws)
    if not table["_columns"] or not table["rows"]:
        return {"_columns": table["_columns"], "rows": []}

    cols = table["_columns"]

    # إيجاد عمود التسمية
    label_idx = None
    for hint in label_col_hints:
        for i, c in enumerate(cols):
            if hint in normalize_col(c):
                label_idx = i
                break
        if label_idx is not None:
            break
    if label_idx is None:
        label_idx = 0

    # إيجاد عمود العدد
    count_idx = None
    for hint in count_col_hints:
        for i, c in enumerate(cols):
            if hint in normalize_col(c):
                count_idx = i
                break
        if count_idx is not None:
            break
    if count_idx is None:
        count_idx = 1 if len(cols) > 1 else 0

    # فرز حسب العدد
    def sort_key(row):
        val = row[count_idx] if count_idx < len(row) else None
        return to_number(val) or 0

    sorted_rows = sorted(table["rows"], key=sort_key, reverse=True)
    return {"_columns": table["_columns"], "rows": sorted_rows[:n]}


# ============================================================
# دالة التحقق الرئيسية
# ============================================================

def validate_excel(wb, book_id: str, filename: str) -> dict:
    """التحقق من امتثال ملف Excel لبروتوكول مرقوم"""
    sheet_names = wb.sheetnames
    report = {
        "book_id": book_id,
        "file": filename,
        "validated_at": datetime.now(timezone.utc).isoformat(),
        "compliance_score": 100,
        "core_sheets": {"found": [], "missing": [], "errors": []},
        "optional_modules": {"found": [], "missing": [], "notes": []},
        "column_warnings": [],
        "data_warnings": [],
        "etl_status": "pending",
    }

    deductions = 0

    # فحص الأوراق الأساسية
    for sheet_name, spec in SCHEMA_REGISTRY["core"].items():
        matched = fuzzy_match_sheet(sheet_names, sheet_name, spec["aliases"])
        if matched:
            report["core_sheets"]["found"].append(matched)
        else:
            report["core_sheets"]["missing"].append(sheet_name)
            deductions += 15

    # فحص الوحدات الاختيارية
    for sheet_name, spec in SCHEMA_REGISTRY["modules"].items():
        matched = fuzzy_match_sheet(sheet_names, sheet_name, spec["aliases"])
        if matched:
            report["optional_modules"]["found"].append(matched)
        else:
            report["optional_modules"]["missing"].append(sheet_name)

    # حساب درجة الامتثال
    report["compliance_score"] = max(0, 100 - deductions)

    return report


# ============================================================
# دالة التحويل الرئيسية
# ============================================================

def convert_excel_to_json(
    excel_path: str,
    book_id: str,
    title: str,
    author: str,
    died_hijri: Optional[int],
    category: str,
    output_dir: str = ".",
) -> tuple[dict, dict]:
    """تحويل Excel إلى view_cache.json + validation_report.json"""

    filename = os.path.basename(excel_path)
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    # حساب SHA256 لاسم الملف (بديل آمن بدون قراءة ثنائية)
    sha256 = hashlib.sha256(filename.encode("utf-8")).hexdigest()[:16]

    # تقرير التحقق
    validation = validate_excel(wb, book_id, filename)

    # ============================================================
    # بناء view_cache.json
    # ============================================================
    cache = {
        "_meta": {
            "schema_version": "1.0.0",
            "book_id": book_id,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "source": "excel",
            "excel_sha256": sha256,
            "generator": "marqoom-etl/1.0",
        },
        "book": {
            "title": title,
            "author": author,
            "died_hijri": died_hijri,
            "category": category,
            "book_id": book_id,
        },
        "summary": {
            "pages": None,
            "parts": None,
            "words_approx": None,
            "chars": None,
            "surahs_detected": None,
        },
        "highlights": {},
        "distributions": {},
        "limits": [],
    }

    def get_sheet(name: str, spec: dict) -> Optional[Any]:
        matched = fuzzy_match_sheet(sheet_names, name, spec.get("aliases", []))
        if matched:
            return wb[matched]
        return None

    # --- لوحة عامة → summary ---
    ws = get_sheet("لوحة عامة", SCHEMA_REGISTRY["core"]["لوحة عامة"])
    if ws:
        cache["summary"] = extract_summary_from_dashboard(ws)

    # --- توزيع الأغراض ---
    ws = get_sheet("توزيع الأغراض", SCHEMA_REGISTRY["core"]["توزيع الأغراض"])
    if ws:
        cache["highlights"]["top_purposes"] = extract_top_n(
            ws, ["غرض", "عبارة", "فئة", "الغرض"], ["عدد", "العدد"], n=15
        )

    # --- الموارد حسب المجال ---
    ws = get_sheet("الموارد حسب المجال", SCHEMA_REGISTRY["core"]["الموارد حسب المجال"])
    if ws:
        cache["highlights"]["top_resources"] = extract_top_n(
            ws, ["مجال", "المجال"], ["عدد", "مؤشر"], n=15
        )

    # --- المؤشرات المركبة ---
    ws = get_sheet("المؤشرات المركبة", SCHEMA_REGISTRY["core"]["المؤشرات المركبة"])
    if ws:
        cache["highlights"]["composite_indicators"] = read_sheet_as_table(ws)

    # --- الكتب المصرح بها ---
    ws = get_sheet("الكتب المصرح بها", SCHEMA_REGISTRY["modules"]["الكتب المصرح بها"])
    if ws:
        cache["highlights"]["top_books"] = extract_top_n(
            ws, ["كتاب", "الكتاب", "نص"], ["عدد", "إحالات"], n=20
        )

    # --- الأعلام ---
    ws = get_sheet("الأعلام", SCHEMA_REGISTRY["modules"]["الأعلام"])
    if ws:
        cache["highlights"]["top_people"] = extract_top_n(
            ws, ["علم", "اسم", "العلم"], ["عدد", "حضور"], n=20
        )

    # --- المصطلحات المنهجية ---
    ws = get_sheet("المصطلحات المنهجية", SCHEMA_REGISTRY["modules"]["المصطلحات المنهجية"])
    if ws:
        cache["highlights"]["terms"] = extract_top_n(
            ws, ["مصطلح", "عبارة", "فئة"], ["عدد"], n=20
        )

    # --- المدارس والاتجاهات ---
    ws = get_sheet("المدارس والاتجاهات", SCHEMA_REGISTRY["modules"]["المدارس والاتجاهات"])
    if ws:
        cache["highlights"]["schools"] = read_sheet_as_table(ws)

    # --- صيغ الأداء والرواية ---
    ws = get_sheet("صيغ الأداء والرواية", SCHEMA_REGISTRY["modules"]["صيغ الأداء والرواية"])
    if ws:
        cache["highlights"]["narration_forms"] = extract_top_n(
            ws, ["صيغة", "الصيغة"], ["عدد"], n=15
        )

    # --- السور (للتفاسير) ---
    ws = get_sheet("السور والصفحات", SCHEMA_REGISTRY["modules"]["السور والصفحات"])
    if ws:
        cache["distributions"]["surahs"] = read_sheet_as_table(ws)

    # --- الكثافة العلمية ---
    ws = get_sheet("الكثافة العلمية", SCHEMA_REGISTRY["modules"]["الكثافة العلمية"])
    if ws:
        cache["distributions"]["density"] = read_sheet_as_table(ws)

    # --- الخلاف والترجيح ---
    ws = get_sheet("الخلاف والترجيح", SCHEMA_REGISTRY["modules"]["الخلاف والترجيح"])
    if ws:
        cache["highlights"]["disagreement"] = read_sheet_as_table(ws)

    # --- ملخص تنفيذي ---
    ws = get_sheet("ملخص تنفيذي", SCHEMA_REGISTRY["modules"]["ملخص تنفيذي"])
    if ws:
        cache["highlights"]["executive_summary"] = read_sheet_as_table(ws)

    # --- حدود الدراسة ---
    ws = get_sheet("حدود الدراسة", SCHEMA_REGISTRY["core"]["حدود الدراسة"])
    if ws:
        table = read_sheet_as_table(ws)
        # تحويل إلى قائمة أزواج [حد، أثر]
        limits = []
        for row in table["rows"]:
            if row and len(row) >= 1:
                label = row[0] if row[0] else ""
                value = row[1] if len(row) > 1 and row[1] else ""
                if label:
                    limits.append([str(label), str(value)])
        cache["limits"] = limits

    # --- للمستفيد السريع ---
    ws = get_sheet("للمستفيد السريع", SCHEMA_REGISTRY["modules"]["للمستفيد السريع"])
    if ws:
        cache["highlights"]["quick_reader"] = read_sheet_as_table(ws)

    # تحديث حالة ETL
    validation["etl_status"] = "success"

    wb.close()

    # حفظ الملفات
    os.makedirs(output_dir, exist_ok=True)
    safe_id = book_id.replace("-", "_")

    cache_path = os.path.join(output_dir, f"{safe_id}_view_cache.json")
    report_path = os.path.join(output_dir, f"{safe_id}_validation_report.json")

    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(validation, f, ensure_ascii=False, indent=2)

    print(f"✅ view_cache.json     → {cache_path}")
    print(f"✅ validation_report   → {report_path}")
    print(f"📊 درجة الامتثال: {validation['compliance_score']}%")
    if validation["core_sheets"]["missing"]:
        print(f"⚠️  أوراق أساسية مفقودة: {validation['core_sheets']['missing']}")

    return cache, validation


# ============================================================
# واجهة سطر الأوامر
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="محوّل مرقوم: Excel → view_cache.json + validation_report.json"
    )
    parser.add_argument("excel", help="مسار ملف Excel")
    parser.add_argument("--book-id", required=True, help="مثال: marqoom-12")
    parser.add_argument("--title", required=True, help="اسم الكتاب")
    parser.add_argument("--author", required=True, help="اسم المؤلف")
    parser.add_argument("--died-hijri", type=int, default=None, help="سنة الوفاة هجرياً")
    parser.add_argument("--category", required=True, help="التصنيف: حديث / فقه / تفسير / عقيدة / تاريخ / لغة")
    parser.add_argument("--output-dir", default=".", help="مجلد الإخراج (افتراضي: المجلد الحالي)")

    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"ERROR: الملف غير موجود: {args.excel}")
        sys.exit(1)

    convert_excel_to_json(
        excel_path=args.excel,
        book_id=args.book_id,
        title=args.title,
        author=args.author,
        died_hijri=args.died_hijri,
        category=args.category,
        output_dir=args.output_dir,
    )


if __name__ == "__main__":
    main()
